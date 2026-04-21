"""
Сводит статистику слов по персонажам из отдельных файлов серий в один xlsx.

Формат источника и вывода описан в Profile — добавление нового сериала
сводится к добавлению записи в PROFILES ниже.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------- стили ----------
FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="305496")
TOTAL_FONT = Font(name=FONT_NAME, bold=True, size=11)
TOTAL_FILL = PatternFill("solid", start_color="FFE699")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_SMALL = Font(name=FONT_NAME, bold=True, size=10)
GRAY_FILL = PatternFill("solid", start_color="F2F2F2")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3864")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass(frozen=True)
class Profile:
    """Описывает формат исходных xlsx-файлов и выходной сводной таблицы."""

    name: str
    sheet_name: str
    # регулярка для извлечения номера серии из имени файла; группа 1 — номер
    episode_pattern: str
    # маркер строки-итога внутри листа (сравнение через `marker in first_cell_upper`)
    total_marker: str
    # индексы колонок (0-based) на листе-источнике
    col_character: int = 0
    col_dialog: int = 1
    col_transcription: int = 2
    col_foreign: int = 3
    col_music: int = 4
    col_burnedin: int = 5
    col_onscreen: int = 6
    col_total: int = 7
    # подписи колонок в выходном файле (порядок = Dialog..Total)
    metric_labels: tuple[str, ...] = (
        "Dialog WC",
        "Transcription WC",
        "Foreign Dialogue",
        "Music And Song",
        "Burnedin Subtitle",
        "Onscreen Text",
        "Total WC",
    )
    character_label: str = "Character"
    episode_label: str = "Episode"


PROFILES: dict[str, Profile] = {
    "default": Profile(
        name="Default",
        sheet_name="Word Count Summary",
        episode_pattern=r"(\d+)\s*СЕРИЯ",
        total_marker="TOTAL WORD COUNT BY TEXT CATEGORY",
    ),
}


# ---------- типы сырых данных ----------
EpisodeRow = tuple  # строка листа (как читает openpyxl)


@dataclass
class EpisodeData:
    number: int
    filename: str
    rows: list[EpisodeRow] = field(default_factory=list)
    total: EpisodeRow | None = None


# ---------- чтение ----------
def collect_episodes(
    files: list[tuple[str, bytes]],
    profile: Profile,
) -> tuple[dict[int, EpisodeData], list[str]]:
    """
    Принимает [(filename, bytes)], возвращает ({ep_num: EpisodeData}, warnings).
    """
    rx = re.compile(profile.episode_pattern)
    episodes: dict[int, EpisodeData] = {}
    warnings: list[str] = []

    for fname, blob in files:
        if fname.startswith("~$") or not fname.lower().endswith(".xlsx"):
            warnings.append(f"{fname}: не xlsx, пропущен")
            continue
        m = rx.search(fname)
        if not m:
            warnings.append(f"{fname}: номер серии не распознан по паттерну, пропущен")
            continue
        ep_num = int(m.group(1))

        wb = load_workbook(io.BytesIO(blob), data_only=True)
        if profile.sheet_name not in wb.sheetnames:
            warnings.append(f"{fname}: нет листа '{profile.sheet_name}', пропущен")
            continue

        ws = wb[profile.sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            warnings.append(f"{fname}: пустой лист, пропущен")
            continue

        data_rows: list[EpisodeRow] = []
        total_row: EpisodeRow | None = None
        min_width = profile.col_total + 1
        for r in rows[1:]:
            if len(r) < min_width:
                r = r + (None,) * (min_width - len(r))
            if not any(v is not None for v in r):
                continue
            first = str(r[profile.col_character] or "").upper()
            if profile.total_marker in first:
                total_row = r
            elif r[profile.col_character] is not None or any(v for v in r[1:]):
                data_rows.append(r)

        if ep_num in episodes:
            warnings.append(
                f"{fname}: серия {ep_num} уже была (из {episodes[ep_num].filename}), "
                f"беру более позднюю"
            )
        episodes[ep_num] = EpisodeData(
            number=ep_num, filename=fname, rows=data_rows, total=total_row
        )

    return episodes, warnings


# ---------- пивот ----------
def build_pivot(
    episodes: dict[int, EpisodeData], profile: Profile
) -> tuple[list[str], dict[str, dict[int, tuple[int, int, int]]]]:
    """character -> {ep -> (dialog, transcription, total)}."""
    all_chars: set[str] = set()
    pivot: dict[str, dict[int, tuple[int, int, int]]] = {}

    for ep, d in episodes.items():
        for r in d.rows:
            cell = r[profile.col_character]
            name = (str(cell).strip() if cell else "") or "(unnamed)"
            all_chars.add(name)
            pivot.setdefault(name, {})[ep] = (
                r[profile.col_dialog] or 0,
                r[profile.col_transcription] or 0,
                r[profile.col_total] or 0,
            )

    return sorted(all_chars), pivot


# ---------- запись листов ----------
def _style_header_row(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def _write_pivot_sheet(
    ws, title: str, metric_index: int, pivot, all_chars, episodes, profile: Profile
) -> None:
    """metric_index: 0=dialog, 2=total (индексы в кортеже pivot[char][ep])."""
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    n_eps = len(episodes)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_eps + 3)
    ws.cell(row=1, column=1).alignment = CENTER

    headers = (
        [profile.character_label]
        + [f"{profile.episode_label} {i}" for i in sorted(episodes.keys())]
        + ["Character Total", "Episodes"]
    )
    _style_header_row(ws, 3, headers)

    totals = {c: sum(v[metric_index] for v in pivot[c].values()) for c in all_chars}
    sorted_chars = sorted(all_chars, key=lambda x: (-totals[x], x))

    row = 4
    first_ep_col = get_column_letter(2)
    last_ep_col = get_column_letter(1 + n_eps)
    sum_col = n_eps + 2

    for name in sorted_chars:
        c = ws.cell(row=row, column=1, value=name)
        c.font, c.alignment, c.border = NORMAL_FONT, LEFT, BORDER
        for i, ep in enumerate(sorted(episodes.keys()), start=2):
            val = pivot[name].get(ep, (0, 0, 0))[metric_index]
            cell = ws.cell(row=row, column=i, value=val if val else None)
            cell.font, cell.alignment, cell.border = NORMAL_FONT, RIGHT, BORDER

        s = ws.cell(row=row, column=sum_col, value=f"=SUM({first_ep_col}{row}:{last_ep_col}{row})")
        s.font, s.alignment, s.border, s.fill = BOLD_SMALL, RIGHT, BORDER, GRAY_FILL

        cnt = ws.cell(
            row=row,
            column=sum_col + 1,
            value=f"=COUNT({first_ep_col}{row}:{last_ep_col}{row})",
        )
        cnt.font, cnt.alignment, cnt.border = NORMAL_FONT, RIGHT, BORDER
        row += 1

    # пустая строка-разделитель: мешает Excel'у захватить EPISODE TOTAL при сортировке
    last_data_row = row - 1
    row += 1

    tc = ws.cell(row=row, column=1, value="EPISODE TOTAL")
    tc.font, tc.fill, tc.alignment, tc.border = TOTAL_FONT, TOTAL_FILL, LEFT, BORDER
    for i in range(2, len(headers) + 1):
        col = get_column_letter(i)
        cell = ws.cell(row=row, column=i, value=f"=SUM({col}4:{col}{last_data_row})")
        cell.font, cell.fill, cell.alignment, cell.border = TOTAL_FONT, TOTAL_FILL, RIGHT, BORDER

    # фильтр на заголовке + данных (без разделителя и строки "EPISODE TOTAL")
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{last_data_row}"

    ws.column_dimensions["A"].width = 32
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "B4"


# ---------- публичный API ----------
def build_workbook_bytes(episodes: dict[int, EpisodeData], profile: Profile) -> bytes:
    all_chars, pivot = build_pivot(episodes, profile)

    wb = Workbook()
    _write_pivot_sheet(
        wb.active,
        "Dialogue Word Count by Character and Episode",
        metric_index=0,  # Dialog WC (колонка B в исходнике — оригинал)
        pivot=pivot,
        all_chars=all_chars,
        episodes=episodes,
        profile=profile,
    )
    wb.active.title = "Dialogue Summary"

    _write_pivot_sheet(
        wb.create_sheet("Transcription Summary"),
        "Transcription Word Count by Character and Episode",
        metric_index=1,  # Transcription WC (колонка C в исходнике — перевод)
        pivot=pivot,
        all_chars=all_chars,
        episodes=episodes,
        profile=profile,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def derive_common_name(filenames: list[str], profile: Profile) -> str:
    """
    Достаёт общее «имя шоу» из набора имён файлов.
    Удаляет расширение и маркер серии (по паттерну профиля), затем оставляет только
    токены, встречающиеся во ВСЕХ именах (сравнение case-insensitive, порядок —
    из первого файла). Возвращает "" если общих слов нет.
    """
    rx = re.compile(profile.episode_pattern, re.IGNORECASE)
    token_sets_upper: list[set[str]] = []
    reference: list[str] = []

    for i, fname in enumerate(filenames):
        stem = Path(fname).stem
        stripped = rx.sub(" ", stem)
        tokens = [t for t in re.split(r"\s+", stripped.strip()) if t]
        if i == 0:
            reference = tokens
        token_sets_upper.append({t.upper() for t in tokens})

    if not reference or not token_sets_upper:
        return ""

    rest = token_sets_upper[1:]
    common = [t for t in reference if all(t.upper() in s for s in rest)]
    return " ".join(common).strip()


def consolidate(
    files: list[tuple[str, bytes]], profile_key: str = "default"
) -> tuple[bytes, dict]:
    """
    Высокоуровневая обёртка для UI.
    Возвращает (xlsx_bytes, info) где info = {'episodes', 'characters', 'warnings', 'common_name'}.
    """
    if profile_key not in PROFILES:
        raise ValueError(f"Неизвестный профиль: {profile_key}")
    profile = PROFILES[profile_key]

    episodes, warnings = collect_episodes(files, profile)
    if not episodes:
        raise ValueError(
            "Не найдено ни одного подходящего файла. "
            "Проверь, что имена содержат номер серии по паттерну профиля."
        )

    xlsx_bytes = build_workbook_bytes(episodes, profile)
    all_chars, _ = build_pivot(episodes, profile)
    # общее имя берём только из файлов, которые реально попали в сводную
    accepted_names = [d.filename for d in episodes.values()]
    common_name = derive_common_name(accepted_names, profile)
    return xlsx_bytes, {
        "episodes": sorted(episodes.keys()),
        "characters": len(all_chars),
        "warnings": warnings,
        "common_name": common_name,
    }


# ---------- CLI (для дебага без UI) ----------
def _cli() -> None:
    import argparse

    p = argparse.ArgumentParser(description="Сводит xlsx-статистику по персонажам.")
    p.add_argument("-i", "--input-dir", default=".")
    p.add_argument("-o", "--output", default=None)
    p.add_argument("-p", "--profile", default="default", choices=list(PROFILES))
    args = p.parse_args()

    folder = Path(args.input_dir).expanduser().resolve()
    files: list[tuple[str, bytes]] = []
    for f in sorted(folder.iterdir()):
        if f.is_file() and f.suffix.lower() == ".xlsx" and not f.name.startswith("~$"):
            files.append((f.name, f.read_bytes()))
    if not files:
        raise SystemExit(f"В папке нет xlsx: {folder}")

    xlsx, info = consolidate(files, args.profile)
    if args.output:
        out = Path(args.output)
    else:
        common = (info.get("common_name") or "").strip()
        out = folder / (f"{common} - Word Count Summary.xlsx" if common else "Word Count Summary.xlsx")
    out.write_bytes(xlsx)
    for w in info["warnings"]:
        print(f"  ! {w}")
    print(f"Серий: {len(info['episodes'])}, персонажей: {info['characters']}")
    print(f"Файл: {out}")


if __name__ == "__main__":
    _cli()
